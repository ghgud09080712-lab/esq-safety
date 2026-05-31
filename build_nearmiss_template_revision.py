from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path(r"C:\Users\zxcas\Desktop\아차사고발굴개선표+체크리스트 위험성평가 양식_개정안.xlsx")


BLUE = "1F4E9E"
LIGHT_BLUE = "D9E2F3"
PALE_BLUE = "EEF3FA"
YELLOW = "FFF2CC"
PALE_GREEN = "E2F0D9"
WHITE = "FFFFFF"
GRAY = "F2F2F2"
TEXT = "1F2937"

thin_blue = Side(style="thin", color="4F6BED")
thin_gray = Side(style="thin", color="BFBFBF")
medium_blue = Side(style="medium", color=BLUE)


def border(color="blue"):
    side = thin_blue if color == "blue" else thin_gray
    return Border(left=side, right=side, top=side, bottom=side)


def fill(color):
    return PatternFill("solid", fgColor=color)


def style_range(ws, cell_range, *, fill_color=None, font=None, align=None, border_style=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill_color:
                cell.fill = fill(fill_color)
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if border_style:
                cell.border = border_style


def merge_label(ws, cell_range, value, fill_color=LIGHT_BLUE, font_size=10):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.fill = fill(fill_color)
    cell.font = Font(name="맑은 고딕", size=font_size, bold=True, color=TEXT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    style_range(ws, cell_range, border_style=border())


def merge_input(ws, cell_range, value=""):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.font = Font(name="맑은 고딕", size=10, color=TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    style_range(ws, cell_range, fill_color=WHITE, border_style=border())


def apply_sheet_base(ws, title_rows=34):
    ws.sheet_view.showGridLines = False
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    for row in range(1, title_rows + 1):
        ws.row_dimensions[row].height = 23
    for col, width in {
        "A": 5,
        "B": 11,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
    }.items():
        ws.column_dimensions[col].width = width


def build_nearmiss_sheet(wb):
    ws = wb.active
    ws.title = "아차사고 발굴·개선표"
    apply_sheet_base(ws, 38)

    ws["A1"] = "[양식7]"
    ws["A1"].font = Font(name="맑은 고딕", size=9, bold=True, color=TEXT)
    ws.merge_cells("A2:H2")
    ws["A2"] = "아차사고 발굴·개선표"
    ws["A2"].font = Font(name="맑은 고딕", size=18, bold=True, color=TEXT)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 32

    for col, title in zip(["F", "G", "H"], ["작 성", "검 토", "승 인"]):
        ws[f"{col}3"] = title
        ws[f"{col}3"].fill = fill(LIGHT_BLUE)
        ws[f"{col}3"].font = Font(name="맑은 고딕", size=10, bold=True)
        ws[f"{col}3"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col}3"].border = border()
        ws[f"{col}4"] = ""
        ws[f"{col}4"].border = border()
        ws.row_dimensions[4].height = 34

    merge_label(ws, "A5:A8", "1. 기본정보")
    merge_label(ws, "B5:B5", "사업장")
    merge_input(ws, "C5:D5", "□ 오영    □ SEM")
    merge_label(ws, "E5:E5", "발굴번호")
    merge_input(ws, "F5:H5")
    merge_label(ws, "B6:B6", "발굴자")
    merge_input(ws, "C6:D6")
    merge_label(ws, "E6:E6", "부서명")
    merge_input(ws, "F6:H6")
    merge_label(ws, "B7:B7", "작성자")
    merge_input(ws, "C7:D7")
    merge_label(ws, "E7:E7", "발굴일시")
    merge_input(ws, "F7:H7", "      년    월    일    시    분경")
    merge_label(ws, "B8:B8", "사고명")
    merge_input(ws, "C8:H8")

    merge_label(ws, "A9:A12", "2. 발생장소")
    merge_label(ws, "B9:B9", "공정")
    merge_input(ws, "C9:D9")
    merge_label(ws, "E9:E9", "장소")
    merge_input(ws, "F9:H9")
    merge_label(ws, "B10:B10", "설비")
    merge_input(ws, "C10:D10")
    merge_label(ws, "E10:E10", "작업명")
    merge_input(ws, "F10:H10")
    merge_label(ws, "B11:B12", "관련 사진")
    merge_input(ws, "C11:H12", "개선 전 사진 또는 현장 사진 첨부")

    merge_label(ws, "A13:A19", "3. 사고개요\n및 원인분석")
    merge_label(ws, "B13:B14", "사고개요")
    merge_input(ws, "C13:H14", "언제, 어디서, 어떤 작업 중, 어떤 상황이 발생했는지 작성")
    merge_label(ws, "B15:B15", "직접원인")
    merge_input(ws, "C15:H15")
    merge_label(ws, "B16:B19", "4M 원인분석")
    for row, label in zip(range(16, 20), ["인적 요인", "설비 요인", "작업환경 요인", "관리적 요인"]):
        merge_label(ws, f"C{row}:C{row}", label, fill_color=PALE_BLUE)
        merge_input(ws, f"D{row}:H{row}")

    merge_label(ws, "A20:A24", "4. 위험요인\n및 예상피해")
    merge_label(ws, "B20:B21", "위험요인")
    merge_input(ws, "C20:H21", "유해·위험요인을 구체적으로 작성")
    merge_label(ws, "B22:B22", "예상피해")
    merge_input(ws, "C22:H22", "예: 끼임, 떨어짐, 화학물질 접촉, 화상 등")
    merge_label(ws, "B23:B23", "위험성")
    merge_input(ws, "C23:H23", "가능성(1~5):      중대성(1~5):      등급: □ 낮음 □ 보통 □ 높음 □ 중대")
    merge_label(ws, "B24:B24", "위험성평가 반영")
    merge_input(ws, "C24:H24", "□ 필요    □ 불필요    □ 체크리스트 위험성평가표로 전환")

    merge_label(ws, "A25:A29", "5. 개선계획")
    merge_label(ws, "B25:B26", "개선조치")
    merge_input(ws, "C25:H26", "위험요인을 제거·대체·통제하기 위한 개선조치 작성")
    merge_label(ws, "B27:B27", "조치구분")
    merge_input(ws, "C27:H27", "□ 즉시조치    □ 임시조치    □ 근본개선    □ 교육/표준개정")
    merge_label(ws, "B28:B28", "담당")
    merge_input(ws, "C28:D28", "담당부서/담당자")
    merge_label(ws, "E28:E28", "개선예정일")
    merge_input(ws, "F28:H28")
    merge_label(ws, "B29:B29", "검토의견")
    merge_input(ws, "C29:H29")

    merge_label(ws, "A30:A34", "6. 완료확인")
    merge_label(ws, "B30:B30", "개선완료일")
    merge_input(ws, "C30:D30")
    merge_label(ws, "E30:E30", "완료확인자")
    merge_input(ws, "F30:H30")
    merge_label(ws, "B31:B32", "개선결과")
    merge_input(ws, "C31:H32", "개선 후 사진, 조치 결과, 작업자 확인내용 작성")
    merge_label(ws, "B33:B33", "효과확인")
    merge_input(ws, "C33:H33", "□ 위험 제거    □ 위험 감소    □ 추가조치 필요")
    merge_label(ws, "B34:B34", "최종상태")
    merge_input(ws, "C34:H34", "□ 완료    □ 보완요청    □ 위험성평가 반영완료")

    note = "※ 본 양식은 아차사고 발굴 → 원인분석 → 개선계획 → 조치 → 검토 → 완료확인 → 위험성평가 반영까지 추적하기 위한 기록입니다."
    ws.merge_cells("A36:H36")
    ws["A36"] = note
    ws["A36"].font = Font(name="맑은 고딕", size=9, color="666666")
    ws["A36"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.print_area = "A1:H36"
    return ws


def build_risk_sheet(wb):
    ws = wb.create_sheet("체크리스트 위험성평가")
    apply_sheet_base(ws, 30)
    for col, width in {
        "A": 4,
        "B": 8,
        "C": 24,
        "D": 11,
        "E": 11,
        "F": 11,
        "G": 24,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
    }.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("B2:F2")
    ws["B2"] = "체크리스트 위험성평가표"
    ws["B2"].font = Font(name="맑은 고딕", size=17, bold=True, color=TEXT)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    merge_label(ws, "G2:G2", "평가공정")
    merge_input(ws, "H2:K2")
    merge_label(ws, "G3:G3", "평가일시")
    merge_input(ws, "H3:K3")
    merge_label(ws, "G4:G4", "평 가 자")
    merge_input(ws, "H4:K4")
    merge_label(ws, "B4:B4", "연계번호")
    merge_input(ws, "C4:F4", "아차사고 발굴번호:")

    ws.merge_cells("B6:C6")
    ws["B6"] = "유해위험요인 조사"
    ws.merge_cells("D6:H6")
    ws["D6"] = "위험성 확인 및 감소대책 수립"
    ws.merge_cells("I6:K6")
    ws["I6"] = "이행확인"
    style_range(ws, "B6:K6", fill_color=LIGHT_BLUE, font=Font(name="맑은 고딕", size=10, bold=True), align=Alignment(horizontal="center", vertical="center"), border_style=border())

    headers = ["번호", "유해위험요인", "적정", "보완", "해당 없음", "위험성 감소대책", "개선 예정일", "개선 완료일", "담당자"]
    ranges = ["B7:B8", "C7:C8", "D7:D8", "E7:E8", "F7:F8", "G7:H8", "I7:I8", "J7:J8", "K7:K8"]
    for rng, header in zip(ranges, headers):
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = header
        cell.fill = fill(PALE_BLUE)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        style_range(ws, rng, border_style=border())

    for row in range(9, 17):
        ws[f"B{row}"] = row - 8
        for col in range(2, 12):
            cell = ws.cell(row, col)
            cell.border = border()
            cell.alignment = Alignment(horizontal="center" if col in [2, 4, 5, 6] else "left", vertical="center", wrap_text=True)
            cell.font = Font(name="맑은 고딕", size=10, color=TEXT)
        ws.row_dimensions[row].height = 34

    merge_label(ws, "B18:B21", "개선대책")
    merge_label(ws, "C18:D18", "관리적 예방대책", fill_color=PALE_GREEN)
    merge_input(ws, "E18:K18", "작업표준 개정, 교육, 표시, 점검주기 변경 등")
    merge_label(ws, "C19:D19", "공학적 예방대책", fill_color=PALE_GREEN)
    merge_input(ws, "E19:K19", "방호장치, 난간, 덮개, 인터록, 설비 개선 등")
    merge_label(ws, "C20:D20", "보호구/비상조치", fill_color=PALE_GREEN)
    merge_input(ws, "E20:K20", "보호구 지급·착용, 비상대응, MSDS, 세척설비 등")
    merge_label(ws, "C21:D21", "최종 확인")
    merge_input(ws, "E21:K21", "□ 조치완료    □ 보완필요    □ 재평가필요")

    dv = DataValidation(type="list", formula1='"적정,보완,해당 없음"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("D9:F16")

    ws.print_area = "B1:K22"
    return ws


def finalize_workbook(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical=cell.alignment.vertical or "center")
        ws.freeze_panes = None


def main():
    wb = Workbook()
    build_nearmiss_sheet(wb)
    build_risk_sheet(wb)
    finalize_workbook(wb)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
