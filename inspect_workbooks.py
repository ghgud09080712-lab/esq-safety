import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def cell_text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def inspect(path):
    wb = load_workbook(path, data_only=False, read_only=False)
    result = {"path": str(path), "sheets": []}
    for ws in wb.worksheets:
        rows = []
        max_row = min(ws.max_row or 0, 80)
        max_col = min(ws.max_column or 0, 40)
        for r in range(1, max_row + 1):
            vals = [cell_text(ws.cell(r, c).value) for c in range(1, max_col + 1)]
            non_empty = [(i + 1, v) for i, v in enumerate(vals) if v]
            if non_empty:
                rows.append({"row": r, "values": non_empty[:30]})
        merged = [str(rng) for rng in list(ws.merged_cells.ranges)[:30]]
        result["sheets"].append(
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "merged": merged,
                "rows": rows[:35],
            }
        )
    return result


for arg in sys.argv[1:]:
    print(json.dumps(inspect(Path(arg)), ensure_ascii=False, indent=2))
