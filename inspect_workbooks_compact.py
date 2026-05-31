import json
import sys
from pathlib import Path

from openpyxl import load_workbook


KEYWORDS = [
    "no",
    "발생",
    "부서",
    "작성",
    "발굴",
    "장소",
    "사고",
    "재해",
    "개선",
    "등급",
    "원인",
    "대책",
    "피해",
    "보고",
    "조사",
    "재발",
]


def text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def row_values(ws, row, max_col):
    vals = []
    for c in range(1, max_col + 1):
        v = text(ws.cell(row, c).value)
        if v:
            vals.append({"col": c, "value": v})
    return vals


def inspect(path):
    wb = load_workbook(path, data_only=False, read_only=False)
    out = {"path": str(path), "sheets": []}
    for ws in wb.worksheets:
        max_col = min(ws.max_column or 0, 45)
        header_candidates = []
        for r in range(1, min(ws.max_row or 0, 120) + 1):
            vals = row_values(ws, r, max_col)
            joined = " ".join(v["value"].lower() for v in vals)
            score = sum(1 for k in KEYWORDS if k in joined)
            if len(vals) >= 4 and score >= 2:
                header_candidates.append({"row": r, "score": score, "values": vals[:35]})
        top_rows = []
        for r in range(1, min(ws.max_row or 0, 20) + 1):
            vals = row_values(ws, r, max_col)
            if vals:
                top_rows.append({"row": r, "values": vals[:35]})
        out["sheets"].append(
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "top_rows": top_rows[:8],
                "header_candidates": header_candidates[:12],
            }
        )
    return out


for arg in sys.argv[1:]:
    print(json.dumps(inspect(Path(arg)), ensure_ascii=False, indent=2))
