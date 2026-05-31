import sys
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\t", " ").strip()


for path_arg in sys.argv[1:]:
    path = Path(path_arg)
    wb = load_workbook(path, data_only=False, read_only=False)
    print(f"FILE\t{path}")
    for ws in wb.worksheets:
        print(f"SHEET\t{ws.title}\tROWS\t{ws.max_row}\tCOLS\t{ws.max_column}")
        max_col = min(ws.max_column or 0, 60)
        printed = 0
        for r in range(1, min(ws.max_row or 0, 80) + 1):
            vals = [text(ws.cell(r, c).value) for c in range(1, max_col + 1)]
            nonempty = [v for v in vals if v]
            if len(nonempty) >= 4:
                print(f"ROW\t{r}\t" + " | ".join(nonempty[:35]))
                printed += 1
            if printed >= 8:
                break
