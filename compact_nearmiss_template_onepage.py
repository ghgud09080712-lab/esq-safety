from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SRC = Path(r"C:\Users\zxcas\Desktop\아차사고발굴개선표+체크리스트 위험성평가 양식_개정안.xlsx")
OUT = Path(r"C:\Users\zxcas\Desktop\아차사고발굴개선표+체크리스트 위험성평가 양식_한장출력안.xlsx")

BLUE = "1F4E9E"
LIGHT_BLUE = "D9E2F3"
PALE_BLUE = "EEF3FA"
WHITE = "FFFFFF"
TEXT = "1F2937"

thin_blue = Side(style="thin", color="4F6BED")


def border():
    return Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)


def fill(color):
    return PatternFill("solid", fgColor=color)


def clear_merged(ws):
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def style_range(ws, cell_range, *, fill_color=None, bold=False, align="center", font_size=9):
    for row in ws[cell_range]:
        for cell in row:
            if fill_color:
                cell.fill = fill(fill_color)
            cell.border = border()
            cell.font = Font(name="맑은 고딕", size=font_size, bold=bold, color=TEXT)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def merge_label(ws, rng, text, fill_color=LIGHT_BLUE, font_size=9):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = text
    style_range(ws, rng, fill_color=fill_color, bold=True, align="center", font_size=font_size)


def merge_input(ws, rng, text="", font_size=9):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = text
    style_range(ws, rng, fill_color=WHITE, bold=False, align="left", font_size=font_size)


def rebuild_nearmiss_sheet(ws):
    clear_merged(ws)
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 4,
        "B": 10,
        "C": 13,
        "D": 13,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 13,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, 28):
        ws.row_dimensions[row].height = 19
    ws.row_dimensions[2].height = 27
    ws.row_dimensions[10].height = 25
    ws.row_dimensions[11].height = 25
    ws.row_dimensions[12].height = 25
    ws.row_dimensions[13].height = 25
    ws.row_dimensions[14].height = 25
    ws.row_dimensions[17].height = 30
    ws.row_dimensions[18].height = 30
    ws.row_dimensions[21].height = 30
    ws.row_dimensions[22].height = 30

    ws["B1"] = "[양식7]"
    ws["B1"].font = Font(name="맑은 고딕", size=8, bold=True)
    ws.merge_cells("B2:F2")
    ws["B2"] = "아차사고 발굴·개선표"
    ws["B2"].font = Font(name="맑은 고딕", size=17, bold=True)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    for col, title in zip(["G", "H", "I"], ["작 성", "검 토", "승 인"]):
        ws[f"{col}2"] = title
        style_range(ws, f"{col}2:{col}2", fill_color=LIGHT_BLUE, bold=True)
        ws[f"{col}3"] = ""
        style_range(ws, f"{col}3:{col}4", fill_color=WHITE)
    ws.merge_cells("G3:G4")
    ws.merge_cells("H3:H4")
    ws.merge_cells("I3:I4")

    merge_label(ws, "B5:B8", "기본\n정보")
    merge_label(ws, "C5:C5", "사업장")
    merge_input(ws, "D5:E5", "□ 오영   □ SEM")
    merge_label(ws, "F5:F5", "발굴번호")
    merge_input(ws, "G5:I5")
    merge_label(ws, "C6:C6", "부서명")
    merge_input(ws, "D6:E6")
    merge_label(ws, "F6:F6", "발굴자")
    merge_input(ws, "G6:I6")
    merge_label(ws, "C7:C7", "작성자")
    merge_input(ws, "D7:E7")
    merge_label(ws, "F7:F7", "발굴일시")
    merge_input(ws, "G7:I7", "년  월  일  시  분경")
    merge_label(ws, "C8:C8", "사고명")
    merge_input(ws, "D8:I8")

    merge_label(ws, "B9:B11", "발생\n상황")
    merge_label(ws, "C9:C9", "공정/장소/설비")
    merge_input(ws, "D9:I9")
    merge_label(ws, "C10:C11", "발생상황")
    merge_input(ws, "D10:I11", "언제, 어디서, 어떤 작업 중, 어떤 상황이 발생했는지 간단히 작성")

    merge_label(ws, "B12:B16", "4M\n원인분석")
    for row, label in zip(range(12, 16), ["Man\n인적", "Machine\n설비", "Media\n환경", "Management\n관리"]):
        merge_label(ws, f"C{row}:C{row}", label, fill_color=PALE_BLUE, font_size=8)
        merge_input(ws, f"D{row}:I{row}", font_size=8)
    merge_label(ws, "C16:C16", "핵심원인")
    merge_input(ws, "D16:I16", "가장 직접적인 원인 1~2개 요약", font_size=8)

    merge_label(ws, "B17:B19", "위험\n평가")
    merge_label(ws, "C17:C17", "위험요인")
    merge_input(ws, "D17:I17")
    merge_label(ws, "C18:C18", "예상피해")
    merge_input(ws, "D18:I18", "예: 끼임, 떨어짐, 화학물질 접촉, 화상 등")
    merge_label(ws, "C19:C19", "위험성")
    merge_input(ws, "D19:I19", "가능성(1~5):     중대성(1~5):     □ 낮음 □ 보통 □ 높음 □ 중대")

    merge_label(ws, "B20:B23", "개선\n및 확인")
    merge_label(ws, "C20:C21", "개선조치")
    merge_input(ws, "D20:I21", "위험요인 제거·대체·통제 방안 작성")
    merge_label(ws, "C22:C22", "담당/기한")
    merge_input(ws, "D22:F22", "담당부서/담당자")
    merge_label(ws, "G22:G22", "예정일")
    merge_input(ws, "H22:I22")
    merge_label(ws, "C23:C23", "완료확인")
    merge_input(ws, "D23:I23", "완료일:        확인자:        □ 완료 □ 보완필요")

    merge_label(ws, "B24:B25", "위험성\n평가연계")
    merge_label(ws, "C24:C24", "반영여부")
    merge_input(ws, "D24:I24", "□ 필요 □ 불필요 □ 체크리스트 위험성평가표 전환")
    merge_label(ws, "C25:C25", "검토의견")
    merge_input(ws, "D25:I25")

    ws.merge_cells("B27:I27")
    ws["B27"] = "※ 4M 원인분석은 핵심 원인을 짧게 작성하고, 세부 위험성평가는 체크리스트 위험성평가표에 연계하여 기록한다."
    ws["B27"].font = Font(name="맑은 고딕", size=8, color="666666")
    ws["B27"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.print_area = "B1:I27"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def tighten_risk_sheet(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_area = "B1:K22"


def main():
    wb = load_workbook(SRC)
    rebuild_nearmiss_sheet(wb["아차사고 발굴·개선표"])
    if "체크리스트 위험성평가" in wb.sheetnames:
      tighten_risk_sheet(wb["체크리스트 위험성평가"])
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
